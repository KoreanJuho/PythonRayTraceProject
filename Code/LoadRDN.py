import openpyxl
from openpyxl.utils import get_column_letter
from typing import List, Dict, Optional, Union
from openpyxl.worksheet.worksheet import Worksheet
from numpy import sin, tan, arcsin, arctan, sqrt
import sys

class LensDataLoader:
    def __init__(self, LDC_path):
        self.LDC_path = LDC_path 
        self.Index_path = r'./Tools/RefractiveIndexData.xlsx'
        self.Index_priority = ["CDGM", "HOYA", "SCHOTT", "HIKARI", "SUMITA", "OHARA"]
        
    def load_rdn(self):
        
        def load_aspheric_coefficients(SurfaceProperty_Sheet: Worksheet, surface_number: int) -> Dict[str, float]:
            # Scan first row of SurfaceProperty sheet to find column with this surface number
            for col in range(2, SurfaceProperty_Sheet.max_column + 1):
                if SurfaceProperty_Sheet[f'{get_column_letter(col)}1'].value == surface_number:
                    # We found the column for this surface's coefficients
                    col_letter = get_column_letter(col)
                    break
            else:
                print(f"No aspheric coefficients found for surface number {surface_number}")
                return {}
            
            # Get aspheric coefficients from the found column
            return {
                SurfaceProperty_Sheet[f'A{coeff_row}'].value: 0 if SurfaceProperty_Sheet[f'{col_letter}{coeff_row}'].value is None else SurfaceProperty_Sheet[f'{col_letter}{coeff_row}'].value
                for coeff_row in range(3, 12)  # 2 to 11
            } , SurfaceProperty_Sheet[f'{col_letter}2'].value
        
        def load_refractive_index(lens_name: str) -> Dict[str, float]:
            # Load refractive index data for a specific lens name from the appropriate excel sheet.
            refractive_index_data: Dict[str, float] = {}

            wb_index = openpyxl.load_workbook(self.Index_path)
            
            # convert lens_name to lower case for comparison
            lens_name_lower = lens_name.lower()
            
            # loop over the sheets in priority order
            for sheet_name in self.Index_priority:
                sheet = wb_index[sheet_name]
                
                # loop over the rows in the sheet, starting from row 3
                for row in range(3, sheet.max_row):
                    lens_name_in_sheet = sheet[f'A{row}'].value
                    if lens_name_in_sheet is not None and lens_name_in_sheet.lower() == lens_name_lower:
                        refractive_index_data = {
                            wavelength: sheet[f'{column}{row}'].value
                            for column, wavelength in zip('BCDEF', ["C", "d", "e", "F", "g"])
                        }
                        return refractive_index_data

            # if we get here, we didn't find the lens in any of the sheets
            raise ValueError(f"Could not find refractive index data for lens '{lens_name}'")
            
        # Load the workbook
        wb = openpyxl.load_workbook(self.LDC_path)

        # Select sheets
        RDN_Sheet: Worksheet = wb['RDN']
        SurfaceProperty_Sheet: Worksheet = wb['SurfaceProperties']

        rdn_data = []

        # The data starts from row 2 and continues until there's a row with no surface type
        row = 2
        while (cell_value := RDN_Sheet[f'A{row}'].value) is not None:
            rdn_name = RDN_Sheet[f'F{row}'].value

            # Create a dictionary for each rdn and append it to the list
            rdn = {
                'Surface_number': row - 2,  # Surface numbering starts from 0
                'Surface_type': RDN_Sheet[f'C{row}'].value,
                'Y_radius': RDN_Sheet[f'D{row}'].value,
                'Thickness': RDN_Sheet[f'E{row}'].value,
                'RefractiveIndex': {wavelength: 1 for wavelength in ["C", "d", "e", "F", "g"]} 
                                if rdn_name is None else load_refractive_index(rdn_name)
            }
            
            # Check if this is the stop surface
            if isinstance(cell_value, str) and cell_value.lower() == 'stop':
                stop_surface = rdn['Surface_number']

            # If surface type is "Asphere", retrieve aspheric coefficients from SurfaceProperty sheet
            if rdn['Surface_type'] == 'Asphere':
                Aspcoeff, k = load_aspheric_coefficients(SurfaceProperty_Sheet, rdn['Surface_number'])
                rdn['K'] = k
                rdn['Aspheric_coefficients'] = Aspcoeff

            rdn_data.append(rdn)
            row += 1

        # The image surface is the last one
        image_surface = rdn_data[-1]['Surface_number']

        # Create a dictionary that contains the fno and the list of rdns
        lens_data = {
            'fno': 2.4,  # Put the actual focal number here
            'yim': 5.25,
            'stop': stop_surface,
            'img' : image_surface,
            'rdn': rdn_data
        }
        
        return lens_data

class ParaxialRayTracing:
    def __init__(self, lens_data):
        self.rdn_data = lens_data['rdn']
        self.stop_surface = lens_data['stop']
        self.image_surface = lens_data['img']
        self.fno = lens_data['fno']
        self.yim = lens_data['yim']

    def calculate_system_matrix(self, start_surface: int, finish_surface: int, wavelength: str) -> Dict[str, float]:
        
        def calc_gaussian_bracket(gaussian_array: List[float], start: int, finish: int) -> float:
            gaussian_bracket = [0] * (finish + 1)
            for current_point in range(start, finish):
                if current_point == start:
                    gaussian_bracket[current_point] = gaussian_array[start]
                elif current_point == start + 1:
                    gaussian_bracket[current_point] = gaussian_array[start] * gaussian_array[start + 1] + 1
                else:
                    gaussian_bracket[current_point] = (gaussian_bracket[current_point - 1] * 
                                                    gaussian_array[current_point] + 
                                                    gaussian_bracket[current_point - 2])
            return gaussian_bracket[finish - 1]
        
        size = finish_surface - start_surface + 1
        ary_len = 2 * size
        
        system_matrix = {
            'A': 0.0,
            'B': 0.0,
            'C': 0.0,
            'D': 0.0,
            'B_dummy': 0.0,
            'D_dummy': 0.0,
        }

        gaussian_array = [-1 * self.rdn_data[0]['Thickness']]
        gaussian_array.extend([(self.rdn_data[start_surface + i // 2]['RefractiveIndex'][wavelength] - self.rdn_data[start_surface - 1 + i // 2]['RefractiveIndex'][wavelength]) / 
                            self.rdn_data[start_surface + i // 2]['Y_radius'] if i % 2 != 0 else 
                            -1 * self.rdn_data[start_surface + i // 2 - 1]['Thickness'] / self.rdn_data[start_surface + i // 2 - 1]['RefractiveIndex'][wavelength]
                            for i in range(1, ary_len)])
        
        system_matrix['A'] = calc_gaussian_bracket(gaussian_array, 1, ary_len - 1)
        system_matrix['B'] = calc_gaussian_bracket(gaussian_array, 0, ary_len - 1)
        system_matrix['C'] = calc_gaussian_bracket(gaussian_array, 1, ary_len)
        system_matrix['D'] = calc_gaussian_bracket(gaussian_array, 0, ary_len)

        gaussian_array[0] = 0
        system_matrix['B_dummy'] = calc_gaussian_bracket(gaussian_array, 0, ary_len - 1)
        system_matrix['D_dummy'] = calc_gaussian_bracket(gaussian_array, 0, ary_len)
        
        return system_matrix
    
    def fio(self, wavelength: str) -> List[Dict[str, float]]:
        system_matrix = self.calculate_system_matrix(1, self.image_surface - 1 , wavelength)
        stop_system_matrix = self.calculate_system_matrix(1, self.stop_surface, wavelength)
        
        v_fio = [{'hmy': 0.0, 'umy': 0.0, 'hcy': 0.0, 'ucy': 0.0} for _ in range(self.image_surface + 1)]
        
        if self.rdn_data[0]['Thickness'] >= 10000000000:
            v_fio[0]['hmy'] = 0.0
            v_fio[0]['umy'] = 0.0
            v_fio[1]['hmy'] = 1 / system_matrix['C'] / 2 / self.fno
        else:
            na = 1 / 2 / self.fno
            nao = - na / system_matrix['D']
            v_fio[0]['hmy'] = 0.0
            v_fio[0]['umy'] = tan(arcsin(nao))
            v_fio[1]['hmy'] = self.rdn_data[0]['Thickness'] * v_fio[0]['umy']

        v_fio[0]['hcy'] = self.yim * system_matrix['D']
        v_fio[0]['ucy'] = stop_system_matrix['A'] / stop_system_matrix['B'] * v_fio[0]['hcy']
        v_fio[1]['hcy'] = stop_system_matrix['B_dummy'] / stop_system_matrix['A'] * v_fio[0]['ucy']
        
        for ray_type in ['my', 'cy']:
            for current_surface in range(1, self.image_surface + 1):
                if current_surface != 1:
                    v_fio[current_surface]['h'+ray_type] = v_fio[current_surface - 1]['h'+ray_type] + self.rdn_data[current_surface - 1]['Thickness'] * v_fio[current_surface - 1]['u'+ray_type]

                v_fio[current_surface]['u'+ray_type] = (self.rdn_data[current_surface - 1]['RefractiveIndex'][wavelength] * v_fio[current_surface - 1]['u'+ray_type] - v_fio[current_surface]['h'+ray_type] * (self.rdn_data[current_surface]['RefractiveIndex'][wavelength] - self.rdn_data[current_surface - 1]['RefractiveIndex'][wavelength]) / self.rdn_data[current_surface]['Y_radius']) / self.rdn_data[current_surface]['RefractiveIndex'][wavelength]

                if current_surface == (self.image_surface - 1) and ray_type == "my":
                    self.rdn_data[current_surface]['Thickness'] = -1 * v_fio[current_surface]['h'+ray_type] / v_fio[current_surface]['u'+ray_type]

        return v_fio
    
    def fir(self, v_fio):
        v_fir = {}
        
        v_fir['EFL'] = v_fio[0]['hmy'] / v_fio[self.image_surface-1]['hcy']
        v_fir['BFL'] = v_fio[self.image_surface-1]['hmy'] / v_fio[self.image_surface-1]['hcy']
        
        v_fir['ENP'] = - v_fio[1]['hcy'] / v_fio[0]['ucy']
        v_fir['EPD'] = 2 * (v_fio[1]['hmy'] + v_fir['ENP'] * v_fio[0]['umy'])
        v_fir['EXP'] = - v_fio[self.image_surface-1]['hcy'] / v_fio[self.image_surface-1]['ucy']
        v_fir['EXD'] = 2 * (v_fio[self.image_surface-1]['hmy'] + v_fir['EXP'] * v_fio[self.image_surface-1]['umy'])
        
        return v_fir

class FiniteRayTracing:
    def __init__(self, lens_data, v_fio, wavelength):
        self.lens_data = lens_data['rdn']
        self.v_fio = v_fio
        self.stop_surface = lens_data['stop']
        self.image_surface = lens_data['img']
        self.wavelength = wavelength
        
    def rsi(self, stp_x, stp_y, ojt_x, ojt_y, wavelength = None):
        if wavelength is None:
            wavelength = self.wavelength
        
        def Calc_Asphere_Surface(current_surface, k, curvature, vtc, v_rsi):
    
            different = None
            counter = 0
            while different is None or (different >= 0.000000000001 or different <= -0.000000000001):
                
                counter += 1
                if counter > 20:
                    print(f"Calc ASP Loop exceeded 20 iterations at surface{current_surface+1}, stopping Code.")
                    sys.exit()
                
                r_square = vtc['X'] ** 2 + vtc['Y'] ** 2
                z_prime = curvature * r_square / (1 + sqrt(1 - (1 + k) * curvature ** 2 * r_square))
                
                time = 2
                for asp in self.lens_data[current_surface + 1]['Aspheric_coefficients'].values():
                    z_prime += asp * r_square ** time
                    time += 1
                
                round_z_over_round_x = curvature * vtc['X'] / sqrt(1 - (1 + k) * curvature ** 2 * r_square)
                round_z_over_round_y = curvature * vtc['Y'] / sqrt(1 - (1 + k) * curvature ** 2 * r_square)
                
                time = 2
                for asp in self.lens_data[current_surface + 1]['Aspheric_coefficients'].values():
                    round_z_over_round_x += 2 * time * asp * r_square ** (time - 1) * vtc['X']
                    round_z_over_round_y += 2 * time * asp * r_square ** (time - 1) * vtc['Y']
                    time += 1
                    
                Vector_Size = sqrt(round_z_over_round_x ** 2 + round_z_over_round_y ** 2 + 1)
                alpha = -1 * round_z_over_round_x / Vector_Size
                beta = -1 * round_z_over_round_y / Vector_Size
                gamma = 1 / Vector_Size

                z = (alpha * v_rsi[current_surface]['L'] + beta * v_rsi[current_surface]['M']) / v_rsi[current_surface]['N'] * vtc['Z'] + gamma * z_prime
                z /= ((alpha * v_rsi[current_surface]['L'] + beta * v_rsi[current_surface]['M']) / v_rsi[current_surface]['N'] + gamma)
                vtc['X'] = v_rsi[current_surface]['L'] / v_rsi[current_surface]['N'] * (z - vtc['Z']) + vtc['X']
                vtc['Y'] = v_rsi[current_surface]['M'] / v_rsi[current_surface]['N'] * (z - vtc['Z']) + vtc['Y']
                vtc['Z'] = z
                
                different = z_prime - vtc['Z']
                
            return vtc, alpha, beta, gamma    
        
        def initial_value(wavelength, stp_x, stp_y, ojt_x, ojt_y):
            vtc = {}
            if ojt_x == 0 and ojt_y == 0:
                vtc['L'] = sin(arctan(self.v_fio[0]['umy'])) * stp_x
                vtc['M'] = sin(arctan(self.v_fio[0]['umy'])) * stp_y
                vtc['N'] = sqrt(1 - vtc['L']**2 - vtc['M']**2)
                vtc['X'] = self.v_fio[1]['hmy'] * stp_x
                vtc['Y'] = self.v_fio[1]['hmy'] * stp_y
                vtc['Z'] = 0
            else:
                vtc['L'] = 0
                vtc['M'] = sin(arctan(self.v_fio[0]['ucy'] * ojt_y)) 
                vtc['N'] = sqrt(1 - vtc['M']**2)
                vtc['X'] = 0
                vtc['Y'] = self.v_fio[0]['hcy'] * ojt_y + vtc['M'] / vtc['N'] * self.lens_data[0]['Thickness']
                vtc['Z'] = 0
                vtc = Newton_Rapson(self.stop_surface, 0, vtc, ojt_y, wavelength)
                vtc['X'] = self.v_fio[1]['hmy'] * stp_x
                vtc['Y'] = vtc['Y'] + self.v_fio[1]['hmy'] * stp_y
                vtc['Z'] = 0
                vtc['L'] = vtc['X'] / sqrt(vtc['X']**2 + (vtc['Y'] - self.v_fio[0]['hcy'] * ojt_y) ** 2 + self.lens_data[0]['Thickness'])
                vtc['M'] = (vtc['Y'] - self.v_fio[0]['hcy'] * ojt_y) / sqrt(vtc['X'] ** 2 + (vtc['Y'] - self.v_fio[0]['hcy'] * ojt_y) ** 2 + self.lens_data[0]['Thickness']**2)
                vtc['N'] = sqrt(1 - vtc['L']**2 - vtc['M']**2)
            return vtc
        
        def Newton_Rapson(target_surface, target, vtc, object_y, wavelength):
            
            counter = 0
            delta_M = 0.00000001
            
            v_rsi = [vtc]
            Raytracing(v_rsi, wavelength)
            
            while True:
                counter += 1
                
                if counter > 20:
                    print(f"Newton Rapson Loop exceeded 20 iterations at surface{target_surface+1}, stopping Code.")
                    sys.exit()
                
                f_xn1 = v_rsi[target_surface]['Y']
                xn1 = v_rsi[0]['Y']

                vtc['L'] = 0
                vtc['M'] = sin(arctan((v_rsi[0]['Y'] - self.v_fio[0]['hcy'] * object_y) / self.lens_data[0]['Thickness']))
                vtc['N'] = sqrt(1 - vtc['L'] ** 2 - vtc['M'] ** 2)
                
                vtc['Y'] = v_rsi[0]['Y'] - delta_M
                
                v_rsi = [vtc]
                Raytracing(v_rsi, wavelength)
                
                f_xn2 = v_rsi[target_surface]['Y']
                f_prime_x = (f_xn1 - f_xn2) / delta_M
                xn2 = xn1 + (target - f_xn1) / f_prime_x
                
                vtc['Y'] = xn2
                vtc['M'] = sin(arctan((vtc['Y'] - self.v_fio[0]['hcy'] * object_y) / self.lens_data[0]['Thickness']))
                vtc['N'] = sqrt(1 - vtc['L'] ** 2 - vtc['M'] ** 2)
                
                v_rsi = [vtc]
                Raytracing(v_rsi, wavelength)

                if target + 0.00000000000001 >= v_rsi[target_surface]['Y'] >= target - 0.00000000000001:
                    return vtc
        
        def Raytracing(v_rsi, wavelength):
            for current_surface in range(self.image_surface):
                vtc = {}
                
                # Transfer Equation
                if current_surface == 0:
                    z = 0
                else:
                    z = v_rsi[current_surface]['Z'] - self.lens_data[current_surface]['Thickness']
                    
                if self.lens_data[current_surface + 1]['Surface_type'] == "Sphere":
                    k = 0
                else:
                    k = self.lens_data[current_surface + 1]['K']
                
                curvature = 1 / self.lens_data[current_surface + 1]['Y_radius']
                F = curvature * (v_rsi[current_surface]['X']**2 + v_rsi[current_surface]['Y']**2 + (1 + k)*z**2) - 2*z
                G = (v_rsi[current_surface]['N'] / self.lens_data[current_surface]['RefractiveIndex'][wavelength]) - curvature * (v_rsi[current_surface]['X'] * (v_rsi[current_surface]['L'] / self.lens_data[current_surface]['RefractiveIndex'][wavelength]) + v_rsi[current_surface]['Y'] * (v_rsi[current_surface]['M'] / self.lens_data[current_surface]['RefractiveIndex'][wavelength]) + z * (v_rsi[current_surface]['N'] / self.lens_data[current_surface]['RefractiveIndex'][wavelength]) * (1 + k))
                
                vtc['LEN'] = F / (G + sqrt(G**2 - curvature * (1 + k * (v_rsi[current_surface]['N'] / self.lens_data[current_surface]['RefractiveIndex'][wavelength])**2) * F))
                
                vtc['X'] = v_rsi[current_surface]['X'] + vtc['LEN'] * (v_rsi[current_surface]['L'] / self.lens_data[current_surface]['RefractiveIndex'][wavelength])
                vtc['Y'] = v_rsi[current_surface]['Y'] + vtc['LEN'] * (v_rsi[current_surface]['M'] / self.lens_data[current_surface]['RefractiveIndex'][wavelength])
                r_square = vtc['X']**2 + vtc['Y']**2
                vtc['Z'] = curvature * r_square / (1 + sqrt(1 - (1 + k) * curvature**2 * r_square))
                
                if self.lens_data[current_surface + 1]['Surface_type'] == 'Sphere':
                    V = sqrt(1 - 2 * curvature * k * vtc['Z'] + curvature**2 * vtc['Z']**2 * k * (1 + k))
                    alpha = -1 * curvature * vtc['X'] / V
                    beta = -1 * curvature * vtc['Y'] / V
                    gamma = (1 - curvature * (1 + k) * vtc['Z']) / V
                else: 
                    vtc, alpha, beta, gamma = Calc_Asphere_Surface(current_surface, k, curvature, vtc, v_rsi)
                
                n_cos_theta = alpha * v_rsi[current_surface]['L'] + beta * v_rsi[current_surface]['M'] + gamma * v_rsi[current_surface]['N']
                n_prime_cos_theta_prime = sqrt(self.lens_data[current_surface + 1]['RefractiveIndex'][wavelength]**2 - self.lens_data[current_surface]['RefractiveIndex'][wavelength]**2 + n_cos_theta**2)

                refractive_power = n_prime_cos_theta_prime - n_cos_theta

                vtc['L'] = v_rsi[current_surface]['L'] + refractive_power * alpha
                vtc['M'] = v_rsi[current_surface]['M'] + refractive_power * beta
                vtc['N'] = v_rsi[current_surface]['N'] + refractive_power * gamma
                
                v_rsi.append(vtc)
                
            return v_rsi
        
        v_rsi = []
        v_rsi.append(initial_value(wavelength, stp_x, stp_y, ojt_x, ojt_y))
        v_rsi = Raytracing(v_rsi, wavelength)
        
        return v_rsi
    
    def Y_SemiAperture(self):
        self.rsi(0, 1 ,0 ,0)

ldl = LensDataLoader(r'./LensDataCenter.xlsx')
lens_data= ldl.load_rdn()

prt = ParaxialRayTracing(lens_data)
value_fio = prt.fio("e")
value_fir = prt.fir(value_fio)

frt = FiniteRayTracing(lens_data, value_fio, "e")
value_rsi = frt.rsi(1, 0.5 ,0 ,0.5)

for a in value_rsi:
    print(a)